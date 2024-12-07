import pandas as pd
import unicodedata
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

class CharacterAligner:
    def __init__(self):
        self.sinonom_similar_dict = {}
        self.quocngu_sinonom_dict = {}

    def load_dictionaries(self):
        try:
            # Đọc file Excel
            sinonom_similar_df = pd.read_excel('dictionary/SinoNom_Similar_Dic_v2.xlsx')
            qn_sinonom_df = pd.read_excel('dictionary/QuocNgu_SinoNom_Dic.xlsx')

            # Xử lý từ điển tương đồng
            for _, row in sinonom_similar_df.iterrows():
                input_char = str(row[sinonom_similar_df.columns[0]]).strip()
                if pd.isna(input_char) or not input_char:
                    continue

                similar_chars_str = str(row[sinonom_similar_df.columns[1]])
                # Tách chuỗi thành list các ký tự, loại bỏ khoảng trắng và ký tự rỗng
                similar_chars = {c.strip() for c in similar_chars_str.split(',') if c.strip()}

                # Thêm chính ký tự đó vào tập tương đồng
                similar_chars.add(input_char)

                # Lưu vào từ điển với key là input_char
                self.sinonom_similar_dict[input_char] = similar_chars

            # Xử lý từ điển Quốc ngữ - Hán Nôm
            for _, row in qn_sinonom_df.iterrows():
                quoc_ngu = str(row['QuocNgu']).strip().lower() if 'QuocNgu' in qn_sinonom_df.columns else str(row[0]).strip().lower()
                han_nom = str(row['SinoNom']).strip() if 'SinoNom' in qn_sinonom_df.columns else str(row[1]).strip()

                if pd.isna(quoc_ngu) or pd.isna(han_nom) or not quoc_ngu or not han_nom:
                    continue

                # Tách từ Quốc ngữ thành các từ đơn
                quoc_ngu_words = quoc_ngu.split()
                for word in quoc_ngu_words:
                    if word not in self.quocngu_sinonom_dict:
                        self.quocngu_sinonom_dict[word] = set()
                    self.quocngu_sinonom_dict[word].add(han_nom)

            # Để debug, in ra một số thống kê
            print(f"Loaded {len(self.sinonom_similar_dict)} characters in similar dictionary")
            print(f"Loaded {len(self.quocngu_sinonom_dict)} words in translation dictionary")

        except Exception as e:
            print(f"Error loading dictionaries: {str(e)}")
            raise

    def get_similar_chars(self, han_nom_char):
        similar_chars = self.sinonom_similar_dict.get(han_nom_char, set())
        similar_chars.add(han_nom_char)
        return similar_chars

    def get_possible_translations(self, quoc_ngu_text):
        quoc_ngu_words = quoc_ngu_text.lower().split()
        possible_translations = set()

        for word in quoc_ngu_words:
            translations = self.quocngu_sinonom_dict.get(word, set())
            if translations:
                if not possible_translations:
                    possible_translations = translations
                else:
                    possible_translations.update(translations)

        return possible_translations

    def analyze_text_pair(self, han_nom_char, quoc_ngu_word):
        # Lấy tập S1 (các chữ tương đồng)
        S1 = self.get_similar_chars(han_nom_char)

        # Lấy tập S2 (các bản dịch có thể)
        S2 = self.get_possible_translations(quoc_ngu_word)

        # Giao của hai tập
        intersection = S1.intersection(S2)

        # Nếu giao của hai tập có nhiều hơn 1 ký tự thì tô xanh
        if len(intersection) > 1:
            return 'blue'
        # Nếu giao của hai tập không có ký tự nào thì tô đỏ
        elif len(intersection) == 0:
            return 'red'
        # Nếu giao của hai tập có đúng 1 ký tự thì không tô màu
        else:
            return None

def main():
    try:
        # Đường dẫn đến file input và output
        input_file = 'MidTerm\\output_csv\\TTVH6_full.xlsx'
        output_file = 'MidTerm\\output_csv\\TTVHVN_6.xlsx'

        # Khởi tạo trình phân tích ký tự
        aligner = CharacterAligner()
        aligner.load_dictionaries()

        # Tải workbook
        wb = load_workbook(input_file)
        sheet = wb.active

        # Xác định các cột
        sinonom_col = None
        quocngu_col = None
        for col in sheet[1]:
            if col.value == 'SinoNom Char':
                sinonom_col = col.column_letter
            elif col.value == 'Âm Hán Việt':
                quocngu_col = col.column_letter

        if not sinonom_col or not quocngu_col:
            raise ValueError("Không tìm thấy cột cần thiết")

        # Xử lý từng hàng trong sheet
        for row in range(2, sheet.max_row + 1):
            han_nom_char = sheet[f'{sinonom_col}{row}'].value
            quoc_ngu_word = sheet[f'{quocngu_col}{row}'].value

            # Phân tích và xác định màu
            color = aligner.analyze_text_pair(han_nom_char, quoc_ngu_word)

            # Tô màu nếu cần
            if color == 'blue':
                sheet[f'{sinonom_col}{row}'].font = Font(color='0000FF')  # Xanh
            elif color == 'red':
                sheet[f'{sinonom_col}{row}'].font = Font(color='FF0000')  # Đỏ)

        # Lưu workbook
        wb.save(output_file)
        print(f"Đã xử lý và lưu file: {output_file}")

    except Exception as e:
        print(f"Có lỗi xảy ra: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()